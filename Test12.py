import openpyxl

def merge_excel_files(file1, file2, output_file):

    workbook1 = openpyxl.load_workbook(file1)
    sheet1 = workbook1.active

    workbook2 = openpyxl.load_workbook(file2)
    sheet2 = workbook2.active

    merged_workbook = openpyxl.Workbook()
    merged_sheet = merged_workbook.active

    header_row = [cell.value for cell in sheet1[1]]
    header_row.extend(["bloodgrp"])  
    merged_sheet.append(header_row)

    for row1 in sheet1.iter_rows(min_row=2, values_only=True):
        name1 = row1[0]

        for row2 in sheet2.iter_rows(min_row=2, values_only=True):
            name2 = row2[0]
            if name1 == name2:
                merged_row = list(row1)
                merged_row.extend([row2[1]])  
                merged_sheet.append(merged_row)
                break

    merged_workbook.save(output_file)
    print(f"Merged data successfully saved to {output_file}.")


if __name__ == "__main__":
    merge_excel_files("file1.xlsx", "file2.xlsx", "output.xlsx")
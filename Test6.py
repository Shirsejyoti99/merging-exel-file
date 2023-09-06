import openpyxl

def merge_columns(file1, file2, file3, output):
    # Load the first workbook
    wb1 = openpyxl.load_workbook(file1)
    sheet1 = wb1.active

    # Load the second workbook
    wb2 = openpyxl.load_workbook(file2)
    sheet2 = wb2.active

    # Load the third workbook
    wb3 = openpyxl.load_workbook(file3)
    sheet3 = wb3.active

    # Create a new workbook for the output
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active

    # Iterate over the columns in the first sheet and append them to the output sheet
    for column in sheet1.iter_cols(values_only=True):
        output_sheet.append(column)

    # Iterate over the columns in the second sheet and append them to the output sheet
    for column in sheet2.iter_cols(values_only=True):
        output_sheet.append(column)

    # Iterate over the columns in the third sheet and append them to the output sheet
    for column in sheet3.iter_cols(values_only=True):
        output_sheet.append(column)

    # Save the output workbook
    output_wb.save(output)
    print("Data merged successfully!")

if __name__ == "__main__":
    merge_columns("file1.xlsx", "file2.xlsx", "file3.xlsx", "output.xlsx")
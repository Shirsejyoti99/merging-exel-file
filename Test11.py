import openpyxl

def merge_excel_files(file1, file2, output_file):
    # Load the first workbook
    workbook1 = openpyxl.load_workbook(file1)
    sheet1 = workbook1.active

    # Load the second workbook
    workbook2 = openpyxl.load_workbook(file2)
    sheet2 = workbook2.active

    # Create a new workbook for the merged data
    merged_workbook = openpyxl.Workbook()
    merged_sheet = merged_workbook.active

    # Write the header row to the merged sheet
    header_row = [cell.value for cell in sheet1[1]]
    merged_sheet.append(header_row)

    # Merge the data based on the 'name' column
    for row1 in sheet1.iter_rows(min_row=2, values_only=True):
        name1 = row1[0]

        # Find the corresponding row in the second sheet
        for row2 in sheet2.iter_rows(min_row=2, values_only=True):
            name2 = row2[0]
            if name1 == name2:
                print(f"Found match for {name1}")
                print(f"row2: {row2}")
                merged_row = list(row1)
                merged_row.extend(list(row2[1:]))
                merged_sheet.append(merged_row)
                break

    # Save the merged workbook
    merged_workbook.save(output_file)
    print(f"Merged data successfully saved to {output_file}.")


if __name__ == "__main__":
    merge_excel_files("file1.xlsx", "file2.xlsx", "output.xlsx")
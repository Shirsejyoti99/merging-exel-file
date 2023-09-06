import openpyxl
def sheet(E1,E2,E3,E4):
    E1_wb= openpyxl.load_workbook(E1)
    E2_wb= openpyxl.load_workbook(E2)
    E3_wb= openpyxl.load_workbook(E3)
    E1_data = E1_wb.active
    E2_data = E2_wb.active
    E3_data = E3_wb.active
    E4_wb = openpyxl.Workbook()
    E4_data = E4_wb.active
    for row in E1_data.iter_rows(values_only=True):
        E4_data.append(row)
    #for row in E2
    for i, row in enumerate(E2_data.iter_rows(values_only=True), start=1):
        for j, value in enumerate(row, start=E2_data.max_column + 1):
            E4_data.cell(row=i, column=j, value=value)

    for i, row in enumerate(E3_data.iter_rows(values_only=True), start=1):
        for j, value in enumerate(row, start=E1_data.max_column + E2_data.max_column + 1):
            E4_data.cell(row=i, column=j, value=value)

    E4_wb.save(E4)
    print("Data merged successfully!")
#if_name_:"_main_":
if __name__ == "__main__":
    sheet("name.xlsx", "blood_group.xlsx","Age.xlsx", "output.xlsx")

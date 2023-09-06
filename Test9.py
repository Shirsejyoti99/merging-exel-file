import openpyxl

def merge_excel_files(file1, file2, file3, output_file):
    workbook1 = openpyxl.load_workbook(file1)
    sheet1 = workbook1.active

    workbook2 = openpyxl.load_workbook(file2)
    sheet2 = workbook2.active

    workbook3 = openpyxl.load_workbook(file3)
    sheet3 = workbook3.active

    merged_workbook = openpyxl.Workbook()
    merged_sheet = merged_workbook.active

    header_row = [cell.value for cell in sheet1[1]]
    header_row.extend(["surname", "bloodgrp"])
    merged_sheet.append(header_row)

    merged_data = {}  # Track merged data using name and surname as key

    for row1 in sheet1.iter_rows(min_row=2, values_only=True):
        name1 = row1[0]
        surname1 = row1[1]
        merged_data[(name1, surname1)] = list(row1)

    for row2 in sheet2.iter_rows(min_row=2, values_only=True):
        name2 = row2[0]
        surname2 = row2[1]
        blood_group = row2[1]

        merged_key = (name2, surname2)
        if merged_key in merged_data:
            merged_data[merged_key].extend([surname2, blood_group])
        else:
            merged_data[merged_key] = [None] * len(header_row)
            merged_data[merged_key][0] = name2
            merged_data[merged_key][1] = surname2
            merged_data[merged_key][-1] = blood_group

    for row3 in sheet3.iter_rows(min_row=2, values_only=True):
        name3 = row3[0]
        surname3 = row3[1]
        additional_info = row3[1]

        merged_key = (name3, surname3)
        if merged_key in merged_data:
            merged_data[merged_key].append(additional_info)
        else:
            merged_data[merged_key] = [None] * len(header_row)
            merged_data[merged_key][0] = name3
            merged_data[merged_key][1] = surname3
            merged_data[merged_key][-2] = additional_info

    for merged_row in merged_data.values():
        if any(merged_row):  # Check if any value exists in the row
            merged_sheet.append(merged_row)

    merged_workbook.save(output_file)
    print(f"Merged data successfully saved to {output_file}.")

if __name__ == "__main__":
    merge_excel_files("file1.xlsx", "file2.xlsx", "file3.xlsx", "output.xlsx")
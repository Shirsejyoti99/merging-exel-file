import openpyxl

def merge_excel_files(file1, file2, output_file):
    # Load the first workbook
    workbook1 = openpyxl.load_workbook(file1)
    sheet1 = workbook1.active

    # Load the second workbook
    workbook2 = openpyxl.load_workbook(file2)
    sheet2 = workbook2.active

    # Create a new workbook for the merged data
    merged_workbook = openpyxl.Workbook()
    merged_sheet = merged_workbook.active

    # Write the header row to the merged sheet
    header_row = [cell.value for cell in sheet1[1]]
    merged_sheet.append(header_row)

    # Create a dictionary to store the age and bloodgrp data
    data_dict = {}
    
    # Extract data from the second sheet and store it in the dictionary
    for row2 in sheet2.iter_rows(min_row=2, values_only=True):
        name2 = row2[0]
        age = row2[1]
        bloodgrp = row2[2]
        data_dict[name2] = (age, bloodgrp) 

    # Merge the data based on the 'name' column
    for row1 in sheet1.iter_rows(min_row=2, values_only=True):
        name1 = row1[0]
        age, bloodgrp = data_dict.get(name1, ("", ""))  # Get the corresponding age and bloodgrp, or empty strings if not found
        merged_row = list(row1)
        merged_row.extend([age, bloodgrp])
        merged_sheet.append(merged_row)

    # Save the merged workbook
    merged_workbook.save(output_file)
    print(f"Merged data successfully saved to {output_file}.")


if __name__ == "__main__":
    merge_excel_files("file1.xlsx", "file2.xlsx", "output.xlsx")
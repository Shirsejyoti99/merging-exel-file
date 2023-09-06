import openpyxl
def sheet(E1,E2,E3,E4):
    E1_wb= openpyxl.load_workbook(E1)
    E2_wb= openpyxl.load_workbook(E2)
    E3_wb= openpyxl.load_workbook(E3)
    E1_data = E1_wb.active
    E2_data = E2_wb.active
    E3_data = E3_wb.active
    E4_wb = openpyxl.Workbook()
    E4_data = E4_wb.active
    for column in E1_data.iter_cols(values_only=True):
        E4_data.append(column)
    for column in E2_data.iter_cols(values_only=True):
        E4_data.append(column)
    for column in E3_data.iter_cols(values_only=True):
        E4_data.append(column)
    E4_wb.save(E4)
    print("Data merged successfully!")
#if_name_:"_main_":
if __name__ == "__main__":
    sheet("name.xlsx", "blood_group.xlsx","Age.xlsx", "output.xlsx")